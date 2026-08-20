"""文件解析服务（阶段 9）。

按 MIME 类型分发解析：文本直接 decode；PDF 用 pypdf；Word 用 python-docx
（段落 + 表格）；Excel 用 openpyxl（遍历单元格）。不支持的类型抛
UnsupportedFileType。解析为同步 CPU 操作，调用方应放入线程池执行。
"""
import zipfile
from io import BytesIO
from xml.etree import ElementTree

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

# 文本类 MIME（直接 decode 为 UTF-8 文本）
TEXT_TYPES = {
    "text/plain",
    "text/markdown",
    "text/x-markdown",  # Firefox 对 .md 上报的 MIME 别名
    "application/json",
    "text/x-python",
    "text/csv",
}

# 支持的文档类 MIME
DOCX_TYPES = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
}
XLSX_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
}

# 图片类 MIME（无内嵌文本层，pypdf 无法提取，交由 OCR 识别）
IMAGE_TYPES = {
    "image/jpeg",
    "image/jpg",
    "image/png",
    "image/bmp",
}

# 全部支持的类型（上传接口校验用）
SUPPORTED_TYPES = TEXT_TYPES | {"application/pdf"} | DOCX_TYPES | XLSX_TYPES | IMAGE_TYPES


class UnsupportedFileType(Exception):
    """不支持的文件类型异常。"""


def is_supported(content_type: str) -> bool:
    """判断 MIME 类型是否支持解析。"""
    return content_type in SUPPORTED_TYPES


# 扩展名（含点，小写）→ 规范 MIME，供上传时浏览器 Content-Type 缺失/不标准时回退识别
_MIME_BY_EXT = {
    ".txt": "text/plain",
    ".md": "text/markdown",
    ".markdown": "text/markdown",
    ".json": "application/json",
    ".py": "text/x-python",
    ".csv": "text/csv",
    ".pdf": "application/pdf",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".bmp": "image/bmp",
}


def coerce_content_type(filename: str | None, content_type: str | None) -> str:
    """规范化上传文件的 MIME 类型。

    浏览器上报的 Content-Type 对部分扩展名不稳定（.md 可能被报成
    application/octet-stream / 空 / text/x-markdown），若上报值已可识别则直接采用，
    否则按扩展名回退识别，保证 .md 等文件能被正确判定。

    Args:
        filename: 原始文件名（用于取扩展名）。
        content_type: 浏览器上报的 Content-Type。

    Returns:
        规范化后的 MIME 类型；无法识别时返回空字符串。
    """
    ct = (content_type or "").strip().lower()
    if ct in SUPPORTED_TYPES:
        return ct
    if filename and "." in filename:
        ext = "." + filename.rsplit(".", 1)[-1].lower()
        if ext in _MIME_BY_EXT:
            return _MIME_BY_EXT[ext]
    return ct


def _parse_text(content: bytes) -> str:
    """文本类型直接解码为 UTF-8 字符串（容错替换非法字节）。"""
    return content.decode("utf-8", errors="replace")


def _parse_pdf(content: bytes) -> str:
    """PDF 用 pypdf 提取全部页面文本。"""
    reader = PdfReader(BytesIO(content))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(pages)


def _parse_docx(content: bytes) -> str:
    """Word 提取段落与表格文本（表格行用制表符分隔、行间换行）。"""
    doc = Document(BytesIO(content))
    parts: list[str] = []
    # 按文档顺序遍历段落与表格，保留原始顺序
    from docx.oxml.ns import qn

    body = doc.element.body
    for child in body.iterchildren():
        if child.tag == qn("w:p"):
            from docx.text.paragraph import Paragraph

            text = Paragraph(child, doc).text
            if text.strip():
                parts.append(text)
        elif child.tag == qn("w:tbl"):
            from docx.table import Table

            for row in Table(child, doc).rows:
                cells = [cell.text.strip() for cell in row.cells]
                if any(cells):
                    parts.append("\t".join(cells))
    return "\n".join(parts)


def _parse_xlsx(content: bytes) -> str:
    """Excel 遍历全部工作表单元格（read_only 模式省内存）。"""
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"【工作表：{sheet.title}】")
            parts.extend(rows)
    wb.close()
    return "\n".join(parts)


def parse_file(content: bytes, content_type: str) -> str:
    """按 MIME 类型分发解析，返回提取的纯文本。

    Args:
        content: 文件二进制内容。
        content_type: MIME 类型。

    Returns:
        提取的纯文本。

    Raises:
        UnsupportedFileType: 不支持的文件类型。
    """
    if content_type in TEXT_TYPES:
        return _parse_text(content)
    if content_type == "application/pdf":
        return _parse_pdf(content)
    if content_type in DOCX_TYPES:
        return _parse_docx(content)
    if content_type in XLSX_TYPES:
        return _parse_xlsx(content)
    if content_type in IMAGE_TYPES:
        return ""  # 图片无内嵌文本层，返回空文本触发 OCR 分支
    raise UnsupportedFileType(f"不支持的文件类型：{content_type}")


def _docx_page_count(content: bytes) -> int | None:
    """尽力读取 docx 页数（读 docProps/app.xml 的 <Pages>）。

    Word 页数由排版引擎计算，python-docx 不做排版故无法精确获取；该字段由
    Word 保存到扩展属性中，可能缺失或不准，拿不到时返回 None。
    """
    try:
        with zipfile.ZipFile(BytesIO(content)) as z:
            if "docProps/app.xml" not in z.namelist():
                return None
            with z.open("docProps/app.xml") as f:
                tree = ElementTree.parse(f)
        ns = {
            "ep": "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
        }
        node = tree.find("ep:Pages", ns)
        if node is not None and node.text:
            return int(node.text)
    except Exception:
        return None
    return None


def extract_file_meta(content: bytes, content_type: str) -> dict:
    """提取文件的结构性元信息（页数 / 工作表数等）。

    Args:
        content: 文件二进制内容。
        content_type: MIME 类型。

    Returns:
        元信息字典：
            PDF  -> {"pages": 页数}
            XLSX -> {"sheets": 工作表数}
            DOCX -> {"pages": 页数或 None}
            文本类 -> {}（无结构性页数概念）
    """
    if content_type == "application/pdf":
        reader = PdfReader(BytesIO(content))
        return {"pages": len(reader.pages)}
    if content_type in XLSX_TYPES:
        wb = load_workbook(BytesIO(content), read_only=True)
        sheets = len(wb.worksheets)
        wb.close()
        return {"sheets": sheets}
    if content_type in DOCX_TYPES:
        return {"pages": _docx_page_count(content)}
    return {}


def _format_size(size: int) -> str:
    """字节数格式化为人类可读（B/KB/MB/GB）。"""
    if size < 1024:
        return f"{size}B"
    if size < 1024 * 1024:
        return f"{size / 1024:.1f}KB"
    if size < 1024 * 1024 * 1024:
        return f"{size / (1024 * 1024):.1f}MB"
    return f"{size / (1024 * 1024 * 1024):.1f}GB"


def build_file_meta_block(
    meta: dict, char_count: int, size: int, content_type: str
) -> str:
    """组装文件元信息块（中文），供压缩后强制拼回正文前。

    Args:
        meta: extract_file_meta 返回的元信息字典。
        char_count: 压缩前原始解析文本的字符数。
        size: 文件字节大小。
        content_type: MIME 类型。

    Returns:
        形如「【文件信息】- 类型/页数/字符数/大小」的纯文本块。
    """
    lines = ["【文件信息】", f"- 类型：{content_type}"]
    if meta.get("pages") is not None:
        lines.append(f"- 页数：{meta['pages']}")
    elif "sheets" in meta:
        lines.append(f"- 工作表数：{meta['sheets']}")
    lines.append(f"- 字符数：{char_count}")
    lines.append(f"- 大小：{_format_size(size)}")
    return "\n".join(lines)
